import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
type(wb)
# print(wb)
# print(wb.sheetnames)
# print(wb['Sheet3'])
# print(wb.active)
sheet = wb['Sheet1']
# print(sheet['A1'])
# print(sheet['A1'].value)
# c = sheet['B1']
# print(c.value)
# print(get_column_letter(c.column))


# print('Row' + str(c.row) + ', Column ' + get_column_letter(c.column) + ' is ' + str(c.value))

# for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2):
#   for cell in row:
#     print(cell)

# for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2):
#   print(row)

# for i in range(1, 8, 2):
#   print(i, sheet.cell(row=i, column=2).value)

# print(get_column_letter(27))
# print(column_index_from_string('AAA'))

# for row_of_cells in sheet['A1:C7']:
#   for cell_object in row_of_cells:
#     print(cell_object.coordinate, cell_object.value)
#   print('---End of Row---')

# for cell_object in sheet['1']:
#   print(cell_object.value)

for row in range(1, 7):
  random_cell_value = sheet['B' + str(row)].value
  print(random_cell_value)


